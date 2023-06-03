from django.shortcuts import render, HttpResponse, redirect
from scApp import models
from openpyxl import load_workbook


def depart_list(request):
    queryset = models.Department.objects.all()
    return render(request, 'depart_list.html', {'queryset': queryset})


def depart_add(request):
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    # 获取post提交内容
    name = request.POST.get('title')
    # 保存数据
    models.Department.objects.create(title=name)
    # 重定向到列表页面
    return redirect("/depart/list/")


def depart_del(request):
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    # return HttpResponse('删除成功')
    return redirect("/depart/list/")


def depart_edit(request, nid):
    if request.method == 'GET':
        row_object = models.Department.objects.filter(id=nid).first()
        # 重定向到列表页面
        return render(request, 'depart_edit.html', {'row_object': row_object})

    # 获取post提交内容
    name = request.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=name)
    return redirect("/depart/list/")


def depart_multi(request):
    # 获取html中input框中的上传的对象，从前面input中name获取
    file_object = request.FILES.get('exc')
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 循环读取excel内容并写入数据库中
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    # cell = sheet.cell(2, 1)
    # print(cell.value)
    return redirect("/depart/list/")
